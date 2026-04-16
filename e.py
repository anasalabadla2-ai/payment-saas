import streamlit as st
import pandas as pd
import pdfplumber
import re
import json
import os
from datetime import datetime, date
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────
#  إعدادات الصفحة
# ─────────────────────────────────────────
st.set_page_config(
    page_title="نظام المطابقة اليومية",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────
#  CSS مخصص
# ─────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;700&display=swap');

* { font-family: 'Cairo', sans-serif !important; }

.main { background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%); }

.header-box {
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
    color: white;
    padding: 25px;
    border-radius: 15px;
    text-align: center;
    margin-bottom: 25px;
    box-shadow: 0 10px 30px rgba(0,0,0,0.3);
}

.header-box h1 { font-size: 2rem; margin: 0; color: #e94560; }
.header-box p  { font-size: 1rem; margin: 5px 0 0; color: #a8b2d8; }

.metric-card {
    background: white;
    padding: 20px;
    border-radius: 12px;
    text-align: center;
    box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    border-top: 4px solid;
    margin-bottom: 10px;
}

.card-green  { border-color: #27ae60; }
.card-blue   { border-color: #2980b9; }
.card-orange { border-color: #e67e22; }
.card-red    { border-color: #e74c3c; }

.metric-value { font-size: 1.8rem; font-weight: 700; }
.metric-label { font-size: 0.85rem; color: #666; margin-top: 5px; }

.match-ok  { background:#d4edda; color:#155724; padding:6px 12px; border-radius:20px; font-weight:600; font-size:.85rem; }
.match-no  { background:#f8d7da; color:#721c24; padding:6px 12px; border-radius:20px; font-weight:600; font-size:.85rem; }
.match-partial { background:#fff3cd; color:#856404; padding:6px 12px; border-radius:20px; font-weight:600; font-size:.85rem; }

.section-title {
    font-size: 1.2rem;
    font-weight: 700;
    color: #1a1a2e;
    padding: 10px 0;
    border-bottom: 3px solid #e94560;
    margin-bottom: 15px;
}

div[data-testid="stButton"] > button {
    background: linear-gradient(135deg, #e94560, #c0392b) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 8px 20px !important;
    font-weight: 600 !important;
    width: 100% !important;
}

div[data-testid="stButton"] > button:hover {
    transform: translateY(-2px);
    box-shadow: 0 5px 15px rgba(233,69,96,0.4) !important;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────
#  ثوابت وملفات التخزين
# ─────────────────────────────────────────
DATA_FILE = "daily_entries.json"
today_str  = date.today().strftime("%Y-%m-%d")

# ─────────────────────────────────────────
#  دوال التخزين
# ─────────────────────────────────────────
def load_data() -> dict:
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_data(data: dict):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def get_today_entries() -> list:
    data = load_data()
    return data.get(today_str, [])

def add_entry(name: str, amount: float):
    data = load_data()
    if today_str not in data:
        data[today_str] = []
    data[today_str].append({
        "id":     len(data[today_str]) + 1,
        "name":   name.strip(),
        "amount": amount,
        "time":   datetime.now().strftime("%H:%M:%S")
    })
    save_data(data)

def delete_entry(entry_id: int):
    data = load_data()
    if today_str in data:
        data[today_str] = [e for e in data[today_str] if e["id"] != entry_id]
        save_data(data)

def clear_today():
    data = load_data()
    data[today_str] = []
    save_data(data)

# ─────────────────────────────────────────
#  استخراج البيانات من الملفات
# ─────────────────────────────────────────
def extract_from_excel(file) -> pd.DataFrame:
    """قراءة ملف Excel واستخراج الأسماء والمبالغ تلقائياً"""
    xl   = pd.ExcelFile(file)
    rows = []

    for sheet in xl.sheet_names:
        df = xl.parse(sheet, header=None)

        # ابحث عن أعمدة الاسم والمبلغ تلقائياً
        name_keywords   = ["اسم", "name", "الاسم", "العميل", "client", "customer", "مستفيد"]
        amount_keywords = ["مبلغ", "amount", "المبلغ", "قيمة", "value", "total", "إجمالي", "رسوم"]

        name_col   = None
        amount_col = None

        for row_idx in range(min(10, len(df))):
            for col_idx in range(len(df.columns)):
                cell = str(df.iloc[row_idx, col_idx]).strip().lower()
                if any(k in cell for k in name_keywords)   and name_col   is None:
                    name_col = col_idx
                if any(k in cell for k in amount_keywords) and amount_col is None:
                    amount_col = col_idx

            if name_col is not None and amount_col is not None:
                # استخدم هذا الصف كرأس
                df.columns = df.iloc[row_idx]
                df = df[row_idx + 1:].reset_index(drop=True)
                break

        # إذا لم نجد رأس واضح، افترض أن العمود الأول اسم والثاني مبلغ
        if name_col is None or amount_col is None:
            df.columns = range(len(df.columns))
            name_col, amount_col = 0, 1

        for _, row in df.iterrows():
            try:
                name   = str(row.iloc[name_col]).strip()
                amount = str(row.iloc[amount_col]).strip()
                if name and name.lower() not in ["nan", "none", "", "اسم", "name"]:
                    clean_amount = re.sub(r"[^\d.]", "", amount)
                    if clean_amount:
                        rows.append({"name": name, "amount": float(clean_amount)})
            except Exception:
                continue

    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=["name", "amount"])


def extract_from_pdf(file) -> pd.DataFrame:
    """استخراج الأسماء والمبالغ من PDF"""
    rows = []
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            # جرب استخراج الجداول أولاً
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    for row in table:
                        if not row:
                            continue
                        text_cells = [str(c).strip() if c else "" for c in row]
                        # ابحث عن خلية تحتوي رقماً
                        for i, cell in enumerate(text_cells):
                            clean = re.sub(r"[^\d.]", "", cell)
                            if clean and float(clean) > 0 and i > 0:
                                name = text_cells[i - 1]
                                if name and name.lower() not in ["nan", "", "none"]:
                                    try:
                                        rows.append({"name": name, "amount": float(clean)})
                                    except Exception:
                                        pass
            else:
                # استخراج نصي
                text = page.extract_text() or ""
                for line in text.split("\n"):
                    # نمط: اسم ثم مبلغ
                    pattern = r"([\u0600-\u06FFa-zA-Z][\u0600-\u06FFa-zA-Z\s]{2,40})\s+([\d,،.]+)"
                    matches = re.findall(pattern, line)
                    for m in matches:
                        name   = m[0].strip()
                        amount = re.sub(r"[,،]", "", m[1])
                        try:
                            rows.append({"name": name, "amount": float(amount)})
                        except Exception:
                            pass

    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=["name", "amount"])


def extract_from_csv(file) -> pd.DataFrame:
    for enc in ["utf-8", "utf-8-sig", "cp1256", "latin-1"]:
        try:
            df = pd.read_csv(file, encoding=enc)
            file.seek(0)
            break
        except Exception:
            file.seek(0)
            df = pd.DataFrame()

    if df.empty:
        return pd.DataFrame(columns=["name", "amount"])

    name_col   = next((c for c in df.columns if any(k in str(c).lower() for k in ["اسم","name","client"])),   df.columns[0])
    amount_col = next((c for c in df.columns if any(k in str(c).lower() for k in ["مبلغ","amount","value"])), df.columns[1] if len(df.columns) > 1 else df.columns[0])

    rows = []
    for _, row in df.iterrows():
        try:
            name   = str(row[name_col]).strip()
            clean  = re.sub(r"[^\d.]", "", str(row[amount_col]))
            if name and clean:
                rows.append({"name": name, "amount": float(clean)})
        except Exception:
            pass
    return pd.DataFrame(rows)


def smart_extract(file, ext: str) -> pd.DataFrame:
    ext = ext.lower().lstrip(".")
    if ext in ["xlsx", "xls"]:
        return extract_from_excel(file)
    elif ext == "pdf":
        return extract_from_pdf(file)
    elif ext == "csv":
        return extract_from_csv(file)
    else:
        # حاول Excel ثم PDF
        try:
            return extract_from_excel(file)
        except Exception:
            file.seek(0)
            return extract_from_pdf(file)

# ─────────────────────────────────────────
#  المطابقة
# ─────────────────────────────────────────
def normalize_name(n: str) -> str:
    n = n.strip().lower()
    n = re.sub(r"\s+", " ", n)
    # توحيد الحروف العربية
    n = n.replace("أ","ا").replace("إ","ا").replace("آ","ا")
    n = n.replace("ة","ه").replace("ى","ي")
    return n

def match_entries(manual: list, file_df: pd.DataFrame, tolerance: float = 0.01) -> pd.DataFrame:
    results = []
    file_df = file_df.copy()
    file_df["_norm_name"]  = file_df["name"].apply(normalize_name)
    file_df["_matched"]    = False

    for entry in manual:
        m_name   = normalize_name(entry["name"])
        m_amount = entry["amount"]

        # بحث بالاسم والمبلغ معاً
        mask = (
            file_df["_norm_name"].apply(lambda x: m_name in x or x in m_name) &
            (abs(file_df["amount"] - m_amount) <= tolerance * max(m_amount, 1))
        )
        match_rows = file_df[mask & ~file_df["_matched"]]

        if not match_rows.empty:
            idx = match_rows.index[0]
            file_df.at[idx, "_matched"] = True
            results.append({
                "الاسم (يدوي)":      entry["name"],
                "المبلغ (يدوي)":     m_amount,
                "الاسم (ملف)":       file_df.at[idx, "name"],
                "المبلغ (ملف)":      file_df.at[idx, "amount"],
                "الفرق":             round(abs(m_amount - file_df.at[idx, "amount"]), 2),
                "الحالة":            "✅ مطابق",
                "_status":           "matched"
            })
        else:
            # بحث بالاسم فقط
            mask_name = file_df["_norm_name"].apply(lambda x: m_name in x or x in m_name)
            name_rows = file_df[mask_name & ~file_df["_matched"]]
            if not name_rows.empty:
                idx = name_rows.index[0]
                file_df.at[idx, "_matched"] = True
                diff = round(m_amount - file_df.at[idx, "amount"], 2)
                results.append({
                    "الاسم (يدوي)":  entry["name"],
                    "المبلغ (يدوي)": m_amount,
                    "الاسم (ملف)":   file_df.at[idx, "name"],
                    "المبلغ (ملف)":  file_df.at[idx, "amount"],
                    "الفرق":         diff,
                    "الحالة":        "⚠️ اسم مطابق / مبلغ مختلف",
                    "_status":       "partial"
                })
            else:
                results.append({
                    "الاسم (يدوي)":  entry["name"],
                    "المبلغ (يدوي)": m_amount,
                    "الاسم (ملف)":   "—",
                    "المبلغ (ملف)":  0,
                    "الفرق":         m_amount,
                    "الحالة":        "❌ غير موجود في الملف",
                    "_status":       "not_found"
                })

    # عناصر في الملف لم تُطابق
    unmatched_file = file_df[~file_df["_matched"]]
    for _, row in unmatched_file.iterrows():
        results.append({
            "الاسم (يدوي)":  "—",
            "المبلغ (يدوي)": 0,
            "الاسم (ملف)":   row["name"],
            "المبلغ (ملف)":  row["amount"],
            "الفرق":         row["amount"],
            "الحالة":        "🔍 موجود في الملف فقط",
            "_status":       "file_only"
        })

    return pd.DataFrame(results)

# ─────────────────────────────────────────
#  تصدير Excel للتقرير
# ─────────────────────────────────────────
def export_report_excel(match_df: pd.DataFrame, manual_entries: list) -> bytes:
    wb = openpyxl.Workbook()

    # ── ورقة المطابقة ──
    ws1 = wb.active
    ws1.title = "تقرير المطابقة"
    ws1.sheet_view.rightToLeft = True

    headers = ["الاسم (يدوي)", "المبلغ (يدوي)", "الاسم (ملف)", "المبلغ (ملف)", "الفرق", "الحالة"]
    colors  = {
        "matched":   "D4EDDA",
        "partial":   "FFF3CD",
        "not_found": "F8D7DA",
        "file_only": "D1ECF1"
    }

    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for ci, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=ci, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    display_cols = [c for c in headers]
    for ri, row in match_df[display_cols].iterrows():
        status = match_df.at[ri, "_status"]
        fill   = PatternFill("solid", fgColor=colors.get(status, "FFFFFF"))
        for ci, val in enumerate(row, 1):
            cell = ws1.cell(row=ri + 2, column=ci, value=val)
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center")
            cell.border    = border

    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 22

    # ── ورقة الإدخال اليومي ──
    ws2 = wb.create_sheet("الإدخال اليومي")
    ws2.sheet_view.rightToLeft = True
    ws2.cell(row=1, column=1, value="م").font  = Font(bold=True)
    ws2.cell(row=1, column=2, value="الاسم").font   = Font(bold=True)
    ws2.cell(row=1, column=3, value="المبلغ").font  = Font(bold=True)
    ws2.cell(row=1, column=4, value="الوقت").font   = Font(bold=True)

    for i, e in enumerate(manual_entries, 2):
        ws2.cell(row=i, column=1, value=e["id"])
        ws2.cell(row=i, column=2, value=e["name"])
        ws2.cell(row=i, column=3, value=e["amount"])
        ws2.cell(row=i, column=4, value=e.get("time", ""))

    # ── ورقة الملخص ──
    ws3 = wb.create_sheet("الملخص")
    ws3.sheet_view.rightToLeft = True
    total_manual = sum(e["amount"] for e in manual_entries)
    matched      = len(match_df[match_df["_status"] == "matched"])
    partial      = len(match_df[match_df["_status"] == "partial"])
    not_found    = len(match_df[match_df["_status"] == "not_found"])
    file_only    = len(match_df[match_df["_status"] == "file_only"])

    summary = [
        ("التاريخ",              today_str),
        ("إجمالي المدخلات",      total_manual),
        ("مطابق تماماً",         matched),
        ("مبلغ مختلف",           partial),
        ("غير موجود في الملف",   not_found),
        ("موجود في الملف فقط",   file_only),
    ]
    for ri, (k, v) in enumerate(summary, 1):
        ws3.cell(row=ri, column=1, value=k).font  = Font(bold=True)
        ws3.cell(row=ri, column=2, value=v)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─────────────────────────────────────────
#  الواجهة الرئيسية
# ─────────────────────────────────────────
st.markdown(f"""
<div class="header-box">
  <h1>📊 نظام المطابقة اليومية</h1>
  <p>أدخل البيانات يدوياً ثم ارفع الكشف لمطابقة تلقائية فورية</p>
  <p style="color:#e94560; font-size:.9rem;">📅 {today_str}</p>
</div>
""", unsafe_allow_html=True)

# ── تهيئة Session State ──
if "file_df"      not in st.session_state: st.session_state.file_df      = None
if "match_result" not in st.session_state: st.session_state.match_result = None
if "refresh"      not in st.session_state: st.session_state.refresh      = 0

entries = get_today_entries()

# ─────────────────────────────────────────
#  الشريط الجانبي – الإدخال اليومي
# ─────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="section-title">➕ إدخال يومي</div>', unsafe_allow_html=True)

    with st.form("entry_form", clear_on_submit=True):
        name   = st.text_input("👤 الاسم", placeholder="أدخل الاسم هنا")
        amount = st.number_input("💰 المبلغ", min_value=0.0, step=0.01, format="%.2f")
        submit = st.form_submit_button("➕ إضافة")

        if submit:
            if name.strip() and amount > 0:
                add_entry(name, amount)
                st.success(f"✅ تمت إضافة: {name} - {amount:,.2f}")
                st.session_state.match_result = None
                st.rerun()
            else:
                st.error("⚠️ يرجى إدخال الاسم والمبلغ")

    st.divider()

    # ── رفع الملف ──
    st.markdown('<div class="section-title">📂 رفع الكشف</div>', unsafe_allow_html=True)
    uploaded = st.file_uploader(
        "Excel / PDF / CSV",
        type=["xlsx", "xls", "pdf", "csv"],
        help="ارفع كشف الحسابات للمطابقة"
    )

    if uploaded:
        with st.spinner("⏳ جاري استخراج البيانات..."):
            ext = uploaded.name.rsplit(".", 1)[-1]
            df  = smart_extract(uploaded, ext)
        if not df.empty:
            st.session_state.file_df = df
            st.success(f"✅ تم استخراج **{len(df)}** سجل من الملف")
        else:
            st.error("❌ لم يتم العثور على بيانات في الملف")

    st.divider()

    # ── تشغيل المطابقة ──
    if st.button("🔄 تشغيل المطابقة"):
        if not entries:
            st.error("لا توجد مدخلات يومية!")
        elif st.session_state.file_df is None:
            st.error("لم يتم رفع أي ملف!")
        else:
            with st.spinner("⏳ جاري المطابقة..."):
                st.session_state.match_result = match_entries(entries, st.session_state.file_df)
            st.success("✅ اكتملت المطابقة!")

    st.divider()

    # ── حذف اليوم ──
    if st.button("🗑️ مسح بيانات اليوم"):
        clear_today()
        st.session_state.match_result = None
        st.session_state.file_df      = None
        st.rerun()

# ─────────────────────────────────────────
#  المحتوى الرئيسي
# ─────────────────────────────────────────
tab1, tab2, tab3 = st.tabs(["📋 المدخلات اليومية", "🔍 نتيجة المطابقة", "📅 السجل التاريخي"])

# ══ تبويب 1: المدخلات اليومية ══
with tab1:
    col1, col2, col3, col4 = st.columns(4)
    total = sum(e["amount"] for e in entries)
    count = len(entries)

    col1.markdown(f'<div class="metric-card card-blue"><div class="metric-value">{count}</div><div class="metric-label">عدد الإدخالات</div></div>', unsafe_allow_html=True)
    col2.markdown(f'<div class="metric-card card-green"><div class="metric-value">{total:,.2f}</div><div class="metric-label">الإجمالي اليومي</div></div>', unsafe_allow_html=True)
    col3.markdown(f'<div class="metric-card card-orange"><div class="metric-value">{(total/count):,.2f if count else 0}</div><div class="metric-label">متوسط المبلغ</div></div>', unsafe_allow_html=True)
    col4.markdown(f'<div class="metric-card card-red"><div class="metric-value">{today_str}</div><div class="metric-label">تاريخ اليوم</div></div>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<div class="section-title">📋 قائمة المدخلات</div>', unsafe_allow_html=True)

    if entries:
        for e in reversed(entries):
            c1, c2, c3, c4 = st.columns([3, 2, 2, 1])
            c1.write(f"**{e['name']}**")
            c2.write(f"💰 {e['amount']:,.2f}")
            c3.write(f"🕐 {e.get('time','')}")
            if c4.button("🗑️", key=f"del_{e['id']}"):
                delete_entry(e["id"])
                st.rerun()
    else:
        st.info("📭 لا توجد مدخلات لليوم. استخدم الشريط الجانبي لإضافة بيانات.")

# ══ تبويب 2: المطابقة ══
with tab2:
    if st.session_state.match_result is not None:
        mr = st.session_state.match_result

        matched   = len(mr[mr["_status"] == "matched"])
        partial   = len(mr[mr["_status"] == "partial"])
        not_found = len(mr[mr["_status"] == "not_found"])
        file_only = len(mr[mr["_status"] == "file_only"])
        total_r   = len(mr)

        pct = round((matched / total_r) * 100, 1) if total_r else 0

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.markdown(f'<div class="metric-card card-green"><div class="metric-value">{matched}</div><div class="metric-label">✅ مطابق</div></div>',   unsafe_allow_html=True)
        c2.markdown(f'<div class="metric-card card-orange"><div class="metric-value">{partial}</div><div class="metric-label">⚠️ مبلغ مختلف</div></div>', unsafe_allow_html=True)
        c3.markdown(f'<div class="metric-card card-red"><div class="metric-value">{not_found}</div><div class="metric-label">❌ غير موجود</div></div>',  unsafe_allow_html=True)
        c4.markdown(f'<div class="metric-card card-blue"><div class="metric-value">{file_only}</div><div class="metric-label">🔍 في الملف فقط</div></div>', unsafe_allow_html=True)
        c5.markdown(f'<div class="metric-card card-green"><div class="metric-value">{pct}%</div><div class="metric-label">نسبة المطابقة</div></div>', unsafe_allow_html=True)

        st.progress(pct / 100)
        st.markdown("---")

        # فلتر
        status_filter = st.multiselect(
            "🔎 فلترة حسب الحالة",
            ["✅ مطابق", "⚠️ اسم مطابق / مبلغ مختلف", "❌ غير موجود في الملف", "🔍 موجود في الملف فقط"],
            default=["✅ مطابق", "⚠️ اسم مطابق / مبلغ مختلف", "❌ غير موجود في الملف", "🔍 موجود في الملف فقط"]
        )

        display_cols = ["الاسم (يدوي)", "المبلغ (يدوي)", "الاسم (ملف)", "المبلغ (ملف)", "الفرق", "الحالة"]
        filtered = mr[mr["الحالة"].apply(lambda x: any(s in x for s in status_filter))][display_cols]
        st.dataframe(filtered, use_container_width=True, hide_index=True)

        # تصدير
        if entries:
            excel_bytes = export_report_excel(mr, entries)
            st.download_button(
                label="📥 تحميل التقرير (Excel)",
                data=excel_bytes,
                file_name=f"تقرير_المطابقة_{today_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.info("⏳ ارفع ملفاً واضغط **تشغيل المطابقة** من الشريط الجانبي.")

# ══ تبويب 3: السجل التاريخي ══
with tab3:
    st.markdown('<div class="section-title">📅 السجل التاريخي</div>', unsafe_allow_html=True)
    all_data = load_data()

    if all_data:
        for day in sorted(all_data.keys(), reverse=True):
            day_entries = all_data[day]
            day_total   = sum(e["amount"] for e in day_entries)

            with st.expander(f"📅 {day}  —  {len(day_entries)} إدخال  |  الإجمالي: {day_total:,.2f}"):
                if day_entries:
                    df_day = pd.DataFrame(day_entries)[["id", "name", "amount", "time"]]
                    df_day.columns = ["م", "الاسم", "المبلغ", "الوقت"]
                    st.dataframe(df_day, use_container_width=True, hide_index=True)
    else:
        st.info("📭 لا يوجد سجل تاريخي بعد.")