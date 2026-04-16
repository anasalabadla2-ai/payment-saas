import streamlit as st
import pandas as pd
import pdfplumber
import re
import json
import os
from datetime import datetime, date
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from fuzzywuzzy import fuzz, process
import warnings
warnings.filterwarnings("ignore")

# ╔══════════════════════════════════════════╗
#   إعدادات الصفحة
# ╚══════════════════════════════════════════╝
st.set_page_config(
    page_title="نظام المطابقة اليومية",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ╔══════════════════════════════════════════╗
#   CSS
# ╚══════════════════════════════════════════╝
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;700&display=swap');
* { font-family: 'Cairo', sans-serif !important; direction: rtl; }

.header-box {
    background: linear-gradient(135deg,#1a1a2e,#16213e,#0f3460);
    color: white; padding: 25px; border-radius: 15px;
    text-align: center; margin-bottom: 25px;
    box-shadow: 0 10px 30px rgba(0,0,0,0.3);
}
.header-box h1 { font-size:2rem; margin:0; color:#e94560; }
.header-box p  { color:#a8b2d8; margin:5px 0 0; }

.metric-card {
    background:white; padding:20px; border-radius:12px;
    text-align:center; box-shadow:0 4px 15px rgba(0,0,0,0.1);
    border-top:4px solid; margin-bottom:10px;
}
.card-green  { border-color:#27ae60; }
.card-blue   { border-color:#2980b9; }
.card-orange { border-color:#e67e22; }
.card-red    { border-color:#e74c3c; }
.metric-value { font-size:1.8rem; font-weight:700; }
.metric-label { font-size:.85rem; color:#666; margin-top:5px; }

.section-title {
    font-size:1.2rem; font-weight:700; color:#1a1a2e;
    padding:10px 0; border-bottom:3px solid #e94560; margin-bottom:15px;
}
</style>
""", unsafe_allow_html=True)

# ╔══════════════════════════════════════════╗
#   ثوابت
# ╚══════════════════════════════════════════╝
DATA_FILE = "daily_entries.json"
today_str = date.today().strftime("%Y-%m-%d")

# ╔══════════════════════════════════════════╗
#   دوال التخزين JSON
# ╚══════════════════════════════════════════╝
def load_data() -> dict:
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_data(data: dict):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def get_today_entries() -> list:
    return load_data().get(today_str, [])

def add_entry(name: str, amount: float):
    data = load_data()
    if today_str not in data:
        data[today_str] = []
    data[today_str].append({
        "id":     len(data[today_str]) + 1,
        "name":   name.strip(),
        "amount": amount,
        "time":   datetime.now().strftime("%H:%M:%S")
    })
    save_data(data)

def delete_entry(entry_id: int):
    data = load_data()
    if today_str in data:
        data[today_str] = [e for e in data[today_str] if e["id"] != entry_id]
        save_data(data)

def clear_today():
    data = load_data()
    data[today_str] = []
    save_data(data)

# ╔══════════════════════════════════════════╗
#   استخراج البيانات من الملفات
# ╚══════════════════════════════════════════╝
NAME_KW   = ["اسم", "name", "الاسم", "العميل", "client", "customer", "مستفيد", "المستفيد"]
AMOUNT_KW = ["مبلغ", "amount", "المبلغ", "قيمة", "value", "total", "إجمالي", "رسوم", "الرسوم"]

def find_col(df_row, keywords: list):
    """ابحث عن أقرب عمود يطابق الكلمات المفتاحية"""
    for cell in df_row:
        cell_str = str(cell).strip().lower()
        if any(k in cell_str for k in keywords):
            return list(df_row).index(cell)
    return None

def clean_amount(val: str) -> float | None:
    val = re.sub(r"[^\d.]", "", str(val))
    try:
        return float(val) if val else None
    except ValueError:
        return None

# ── Excel ──
def extract_excel(file) -> pd.DataFrame:
    rows = []
    xl = pd.ExcelFile(file)
    for sheet in xl.sheet_names:
        raw = xl.parse(sheet, header=None)
        name_col = amount_col = header_row = None

        for ri in range(min(15, len(raw))):
            row = raw.iloc[ri]
            nc = find_col(row, NAME_KW)
            ac = find_col(row, AMOUNT_KW)
            if nc is not None and ac is not None:
                name_col, amount_col, header_row = nc, ac, ri
                break

        # لو ما فيه header واضح، خذ أول عمودين
        if name_col is None:
            name_col, amount_col, header_row = 0, 1, -1

        start = (header_row + 1) if header_row is not None else 0
        for ri in range(start, len(raw)):
            row = raw.iloc[ri]
            try:
                name = str(row.iloc[name_col]).strip()
                amt  = clean_amount(str(row.iloc[amount_col]))
                if name and name.lower() not in ["nan","none",""] and amt and amt > 0:
                    rows.append({"name": name, "amount": amt})
            except Exception:
                continue
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=["name","amount"])

# ── PDF ──
def extract_pdf(file) -> pd.DataFrame:
    rows = []
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    # حدد أعمدة الاسم والمبلغ من الصف الأول
                    if not table or not table[0]:
                        continue
                    header = [str(c).strip().lower() if c else "" for c in table[0]]
                    nc = next((i for i,h in enumerate(header) if any(k in h for k in NAME_KW)), None)
                    ac = next((i for i,h in enumerate(header) if any(k in h for k in AMOUNT_KW)), None)
                    start = 1 if (nc is not None and ac is not None) else 0
                    if nc is None: nc = 0
                    if ac is None: ac = 1
                    for row in table[start:]:
                        if not row or len(row) <= max(nc, ac):
                            continue
                        name = str(row[nc]).strip() if row[nc] else ""
                        amt  = clean_amount(str(row[ac]) if row[ac] else "")
                        if name and name.lower() not in ["nan","none",""] and amt and amt > 0:
                            rows.append({"name": name, "amount": amt})
            else:
                # استخراج نصي بـ regex
                text = page.extract_text() or ""
                for line in text.split("\n"):
                    m = re.search(
                        r'([\u0600-\u06FFa-zA-Z][\u0600-\u06FFa-zA-Z\s]{2,40})\s+([\d,،.]+)',
                        line
                    )
                    if m:
                        name = m.group(1).strip()
                        amt  = clean_amount(m.group(2).replace("،",",").replace(",",""))
                        if amt and amt > 0:
                            rows.append({"name": name, "amount": amt})
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=["name","amount"])

# ── CSV ──
def extract_csv(file) -> pd.DataFrame:
    df = pd.DataFrame()
    for enc in ["utf-8","utf-8-sig","cp1256","latin-1"]:
        try:
            df = pd.read_csv(file, encoding=enc)
            file.seek(0)
            break
        except Exception:
            file.seek(0)

    if df.empty:
        return pd.DataFrame(columns=["name","amount"])

    nc = next((c for c in df.columns if any(k in str(c).lower() for k in NAME_KW)),   df.columns[0])
    ac = next((c for c in df.columns if any(k in str(c).lower() for k in AMOUNT_KW)), df.columns[min(1,len(df.columns)-1)])

    rows = []
    for _, row in df.iterrows():
        name = str(row[nc]).strip()
        amt  = clean_amount(str(row[ac]))
        if name and name.lower() not in ["nan","none",""] and amt and amt > 0:
            rows.append({"name": name, "amount": amt})
    return pd.DataFrame(rows)

# ── موزع ذكي ──
def smart_extract(file, filename: str) -> pd.DataFrame:
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext in ["xlsx","xls"]:
        return extract_excel(file)
    elif ext == "pdf":
        return extract_pdf(file)
    elif ext == "csv":
        return extract_csv(file)
    else:
        # جرب Excel ثم PDF
        try:
            result = extract_excel(file)
            if not result.empty:
                return result
        except Exception:
            pass
        file.seek(0)
        return extract_pdf(file)

# ╔══════════════════════════════════════════╗
#   المطابقة بـ FuzzyWuzzy
# ╚══════════════════════════════════════════╝
def normalize(text: str) -> str:
    text = text.strip().lower()
    text = re.sub(r"\s+", " ", text)
    # توحيد الحروف العربية
    for a, b in [("أ","ا"),("إ","ا"),("آ","ا"),("ة","ه"),("ى","ي")]:
        text = text.replace(a, b)
    return text

def match_entries(
    manual: list,
    file_df: pd.DataFrame,
    name_threshold: int = 80,
    amount_tolerance: float = 0.01
) -> pd.DataFrame:
    """
    مطابقة ذكية باستخدام FuzzyWuzzy:
    - إذا تطابق الاسم (≥ threshold) والمبلغ → ✅ مطابق
    - إذا تطابق الاسم فقط               → ⚠️ مبلغ مختلف
    - إذا لم يوجد الاسم أصلاً            → ❌ غير موجود
    - ما في الملف ولم يُطابق             → 🔍 في الملف فقط
    """
    results    = []
    file_copy  = file_df.copy().reset_index(drop=True)
    file_copy["_norm"] = file_copy["name"].apply(normalize)
    file_copy["_used"] = False

    file_names_norm = file_copy["_norm"].tolist()

    for entry in manual:
        m_name   = normalize(entry["name"])
        m_amount = entry["amount"]

        # أفضل تطابق للاسم باستخدام FuzzyWuzzy
        match_result = process.extractOne(
            m_name,
            file_names_norm,
            scorer=fuzz.token_sort_ratio
        )

        if match_result and match_result[1] >= name_threshold:
            best_name  = match_result[0]
            best_score = match_result[1]
            # ابحث عن الصف غير المستخدم
            candidates = file_copy[
                (file_copy["_norm"] == best_name) &
                (~file_copy["_used"])
            ]

            if candidates.empty:
                # خذ أي صف بهذا الاسم
                candidates = file_copy[file_copy["_norm"] == best_name]

            if not candidates.empty:
                idx      = candidates.index[0]
                file_amt = file_copy.at[idx, "amount"]
                diff     = round(m_amount - file_amt, 2)
                amt_ok   = abs(diff) <= amount_tolerance * max(m_amount, 1)

                file_copy.at[idx, "_used"] = True

                if amt_ok:
                    status     = "✅ مطابق"
                    status_key = "matched"
                else:
                    status     = "⚠️ اسم مطابق / مبلغ مختلف"
                    status_key = "partial"

                results.append({
                    "الاسم (يدوي)":      entry["name"],
                    "المبلغ (يدوي)":     m_amount,
                    "الاسم (ملف)":       file_copy.at[idx, "name"],
                    "المبلغ (ملف)":      file_amt,
                    "الفرق":             diff,
                    "درجة التشابه":      f"{best_score}%",
                    "الحالة":            status,
                    "_status":           status_key
                })
            else:
                results.append(_not_found_row(entry))
        else:
            results.append(_not_found_row(entry))

    # ما تبقى في الملف
    for idx, row in file_copy[~file_copy["_used"]].iterrows():
        results.append({
            "الاسم (يدوي)":  "—",
            "المبلغ (يدوي)": 0,
            "الاسم (ملف)":   row["name"],
            "المبلغ (ملف)":  row["amount"],
            "الفرق":         row["amount"],
            "درجة التشابه":  "—",
            "الحالة":        "🔍 موجود في الملف فقط",
            "_status":       "file_only"
        })

    return pd.DataFrame(results)

def _not_found_row(entry: dict) -> dict:
    return {
        "الاسم (يدوي)":  entry["name"],
        "المبلغ (يدوي)": entry["amount"],
        "الاسم (ملف)":   "—",
        "المبلغ (ملف)":  0,
        "الفرق":         entry["amount"],
        "درجة التشابه":  "0%",
        "الحالة":        "❌ غير موجود في الملف",
        "_status":       "not_found"
    }

# ╔══════════════════════════════════════════╗
#   تصدير Excel
# ╚══════════════════════════════════════════╝
def export_excel(match_df: pd.DataFrame, manual: list) -> bytes:
    wb = openpyxl.Workbook()

    STATUS_COLORS = {
        "matched":   "D4EDDA",
        "partial":   "FFF3CD",
        "not_found": "F8D7DA",
        "file_only": "D1ECF1"
    }
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill("solid", fgColor="1A1A2E")
    h_font = Font(color="FFFFFF", bold=True, size=11)

    # ── ورقة 1: المطابقة ──
    ws1 = wb.active
    ws1.title = "تقرير المطابقة"
    ws1.sheet_view.rightToLeft = True

    cols = ["الاسم (يدوي)","المبلغ (يدوي)","الاسم (ملف)","المبلغ (ملف)","الفرق","درجة التشابه","الحالة"]
    for ci, h in enumerate(cols, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.fill = h_fill; c.font = h_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for ri, row in match_df[cols].iterrows():
        color = STATUS_COLORS.get(match_df.at[ri, "_status"], "FFFFFF")
        fill  = PatternFill("solid", fgColor=color)
        for ci, val in enumerate(row, 1):
            c = ws1.cell(row=ri+2, column=ci, value=val)
            c.fill = fill
            c.alignment = Alignment(horizontal="center")
            c.border = border

    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 24

    # ── ورقة 2: الإدخال اليومي ──
    ws2 = wb.create_sheet("الإدخال اليومي")
    ws2.sheet_view.rightToLeft = True
    for ci, h in enumerate(["م","الاسم","المبلغ","الوقت"], 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True); c.fill = h_fill; c.font = h_font
    for i, e in enumerate(manual, 2):
        ws2.cell(row=i, column=1, value=e["id"])
        ws2.cell(row=i, column=2, value=e["name"])
        ws2.cell(row=i, column=3, value=e["amount"])
        ws2.cell(row=i, column=4, value=e.get("time",""))

    # ── ورقة 3: الملخص ──
    ws3 = wb.create_sheet("الملخص")
    ws3.sheet_view.rightToLeft = True
    total_manual = sum(e["amount"] for e in manual)
    stats = [
        ("التاريخ",            today_str),
        ("إجمالي المدخلات اليدوية", total_manual),
        ("✅ مطابق تماماً",    len(match_df[match_df["_status"]=="matched"])),
        ("⚠️ مبلغ مختلف",     len(match_df[match_df["_status"]=="partial"])),
        ("❌ غير موجود",       len(match_df[match_df["_status"]=="not_found"])),
        ("🔍 في الملف فقط",    len(match_df[match_df["_status"]=="file_only"])),
    ]
    for ri, (k, v) in enumerate(stats, 1):
        ws3.cell(row=ri, column=1, value=k).font = Font(bold=True)
        ws3.cell(row=ri, column=2, value=v)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ╔══════════════════════════════════════════╗
#   الواجهة الرئيسية
# ╚══════════════════════════════════════════╝
st.markdown(f"""
<div class="header-box">
  <h1>📊 نظام المطابقة اليومية</h1>
  <p>أدخل البيانات يدوياً ثم ارفع الكشف للمطابقة الذكية الفورية</p>
  <p style="color:#e94560;">📅 {today_str}</p>
</div>
""", unsafe_allow_html=True)

# Session State
for key in ["file_df","match_result"]:
    if key not in st.session_state:
        st.session_state[key] = None

entries = get_today_entries()

# ╔══════════════════════════════════════════╗
#   الشريط الجانبي
# ╚══════════════════════════════════════════╝
with st.sidebar:
    st.markdown('<div class="section-title">➕ إدخال يومي</div>', unsafe_allow_html=True)

    with st.form("entry_form", clear_on_submit=True):
        name   = st.text_input("👤 الاسم", placeholder="أدخل الاسم")
        amount = st.number_input("💰 المبلغ", min_value=0.0, step=0.01, format="%.2f")
        if st.form_submit_button("➕ إضافة"):
            if name.strip() and amount > 0:
                add_entry(name, amount)
                st.success(f"✅ {name} — {amount:,.2f}")
                st.session_state.match_result = None
                st.rerun()
            else:
                st.error("⚠️ أدخل الاسم والمبلغ")

    st.divider()

    st.markdown('<div class="section-title">📂 رفع الكشف</div>', unsafe_allow_html=True)
    uploaded = st.file_uploader(
        "Excel / PDF / CSV",
        type=["xlsx","xls","pdf","csv"],
        help="ارفع كشف الحسابات"
    )
    if uploaded:
        with st.spinner("⏳ استخراج البيانات..."):
            df = smart_extract(uploaded, uploaded.name)
        if not df.empty:
            st.session_state.file_df = df
            st.success(f"✅ {len(df)} سجل تم استخراجه")
            with st.expander("👁️ معاينة الملف"):
                st.dataframe(df.head(10), use_container_width=True)
        else:
            st.error("❌ لم يُعثر على بيانات في الملف")

    st.divider()

    # ── إعدادات المطابقة ──
    st.markdown('<div class="section-title">⚙️ إعدادات المطابقة</div>', unsafe_allow_html=True)
    threshold = st.slider("🎯 حساسية تشابه الاسم", 50, 100, 80,
                          help="كلما زادت القيمة كان التطابق أدق")
    tolerance = st.slider("💱 هامش فرق المبلغ %", 0, 10, 1) / 100

    st.divider()

    col_run, col_clr = st.columns(2)
    with col_run:
        if st.button("🔄 مطابقة"):
            if not entries:
                st.error("لا يوجد مدخلات!")
            elif st.session_state.file_df is None:
                st.error("لم يُرفع ملف!")
            else:
                with st.spinner("⏳ جاري المطابقة..."):
                    st.session_state.match_result = match_entries(
                        entries,
                        st.session_state.file_df,
                        name_threshold=threshold,
                        amount_tolerance=tolerance
                    )
                st.success("✅ اكتملت!")
    with col_clr:
        if st.button("🗑️ مسح"):
            clear_today()
            st.session_state.match_result = None
            st.session_state.file_df      = None
            st.rerun()

# ╔══════════════════════════════════════════╗
#   المحتوى — تبويبات
# ╚══════════════════════════════════════════╝
tab1, tab2, tab3 = st.tabs(["📋 المدخلات اليومية", "🔍 نتائج المطابقة", "📅 السجل التاريخي"])

# ═══════════ تبويب 1 ═══════════
with tab1:
    entries = get_today_entries()
    total   = sum(e["amount"] for e in entries)
    count   = len(entries)
    avg     = total / count if count else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(f'<div class="metric-card card-blue"><div class="metric-value">{count}</div><div class="metric-label">عدد الإدخالات</div></div>', unsafe_allow_html=True)
    c2.markdown(f'<div class="metric-card card-green"><div class="metric-value">{total:,.2f}</div><div class="metric-label">الإجمالي</div></div>', unsafe_allow_html=True)
    c3.markdown(f'<div class="metric-card card-orange"><div class="metric-value">{avg:,.2f}</div><div class="metric-label">المتوسط</div></div>', unsafe_allow_html=True)
    c4.markdown(f'<div class="metric-card card-red"><div class="metric-value">{today_str}</div><div class="metric-label">التاريخ</div></div>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<div class="section-title">📋 قائمة اليوم</div>', unsafe_allow_html=True)

    if entries:
        # عرض جدول
        df_show = pd.DataFrame(entries)[["id","name","amount","time"]]
        df_show.columns = ["م","الاسم","المبلغ","الوقت"]
        st.dataframe(df_show, use_container_width=True, hide_index=True)

        # حذف فردي
        st.markdown("**حذف إدخال:**")
        del_id = st.number_input("رقم الإدخال للحذف", min_value=1, step=1)
        if st.button("🗑️ حذف هذا الإدخال"):
            delete_entry(int(del_id))
            st.rerun()
    else:
        st.info("📭 لا توجد مدخلات اليوم — أضف من الشريط الجانبي")

# ═══════════ تبويب 2 ═══════════
with tab2:
    if st.session_state.match_result is not None:
        mr = st.session_state.match_result
        matched   = len(mr[mr["_status"]=="matched"])
        partial   = len(mr[mr["_status"]=="partial"])
        not_found = len(mr[mr["_status"]=="not_found"])
        file_only = len(mr[mr["_status"]=="file_only"])
        total_r   = len(mr)
        pct       = round(matched/total_r*100, 1) if total_r else 0

        c1,c2,c3,c4,c5 = st.columns(5)
        c1.markdown(f'<div class="metric-card card-green"><div class="metric-value">{matched}</div><div class="metric-label">✅ مطابق</div></div>',    unsafe_allow_html=True)
        c2.markdown(f'<div class="metric-card card-orange"><div class="metric-value">{partial}</div><div class="metric-label">⚠️ فرق مبلغ</div></div>', unsafe_allow_html=True)
        c3.markdown(f'<div class="metric-card card-red"><div class="metric-value">{not_found}</div><div class="metric-label">❌ غير موجود</div></div>',  unsafe_allow_html=True)
        c4.markdown(f'<div class="metric-card card-blue"><div class="metric-value">{file_only}</div><div class="metric-label">🔍 ملف فقط</div></div>',   unsafe_allow_html=True)
        c5.markdown(f'<div class="metric-card card-green"><div class="metric-value">{pct}%</div><div class="metric-label">نسبة المطابقة</div></div>',    unsafe_allow_html=True)

        st.progress(pct / 100)
        st.markdown("---")

        # فلتر
        filter_opts = st.multiselect(
            "🔎 فلترة",
            ["✅ مطابق","⚠️ اسم مطابق / مبلغ مختلف","❌ غير موجود في الملف","🔍 موجود في الملف فقط"],
            default=["✅ مطابق","⚠️ اسم مطابق / مبلغ مختلف","❌ غير موجود في الملف","🔍 موجود في الملف فقط"]
        )

        display_cols = ["الاسم (يدوي)","المبلغ (يدوي)","الاسم (ملف)","المبلغ (ملف)","الفرق","درجة التشابه","الحالة"]
        filtered = mr[mr["الحالة"].apply(lambda x: any(s in x for s in filter_opts))]
        st.dataframe(filtered[display_cols], use_container_width=True, hide_index=True)

        # تحميل
        entries_now = get_today_entries()
        if entries_now:
            xl_bytes = export_excel(mr, entries_now)
            st.download_button(
                label="📥 تحميل التقرير Excel",
                data=xl_bytes,
                file_name=f"تقرير_مطابقة_{today_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.info("⏳ ارفع ملفاً واضغط **🔄 مطابقة** من الشريط الجانبي")

# ═══════════ تبويب 3 ═══════════
with tab3:
    st.markdown('<div class="section-title">📅 السجل التاريخي</div>', unsafe_allow_html=True)
    all_data = load_data()
    if all_data:
        for day in sorted(all_data.keys(), reverse=True):
            day_ent   = all_data[day]
            day_total = sum(e["amount"] for e in day_ent)
            with st.expander(f"📅 {day}  —  {len(day_ent)} إدخال  |  الإجمالي: {day_total:,.2f}"):
                if day_ent:
                    df_d = pd.DataFrame(day_ent)[["id","name","amount","time"]]
                    df_d.columns = ["م","الاسم","المبلغ","الوقت"]
                    st.dataframe(df_d, use_container_width=True, hide_index=True)
    else:
        st.info("📭 لا يوجد سجل بعد")