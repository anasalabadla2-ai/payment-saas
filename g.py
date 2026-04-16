import streamlit as st
import pandas as pd
import pdfplumber
import re
import json
import os
from datetime import datetime, date
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from fuzzywuzzy import fuzz, process
import warnings
warnings.filterwarnings("ignore")

# ╔══════════════════════════════════════════╗
#   إعدادات الصفحة
# ╚══════════════════════════════════════════╝
st.set_page_config(
    page_title="نظام المطابقة اليومية",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;700&display=swap');
* { font-family: 'Cairo', sans-serif !important; }
.header-box {
    background: linear-gradient(135deg,#1a1a2e,#16213e,#0f3460);
    color:white; padding:25px; border-radius:15px;
    text-align:center; margin-bottom:25px;
    box-shadow:0 10px 30px rgba(0,0,0,0.3);
}
.header-box h1 { font-size:2rem; margin:0; color:#e94560; }
.header-box p  { color:#a8b2d8; margin:5px 0 0; }
.metric-card {
    background:white; padding:20px; border-radius:12px;
    text-align:center; box-shadow:0 4px 15px rgba(0,0,0,0.1);
    border-top:4px solid; margin-bottom:10px;
}
.card-green  { border-color:#27ae60; }
.card-blue   { border-color:#2980b9; }
.card-orange { border-color:#e67e22; }
.card-red    { border-color:#e74c3c; }
.metric-value { font-size:1.8rem; font-weight:700; }
.metric-label { font-size:.85rem; color:#666; margin-top:5px; }
.section-title {
    font-size:1.2rem; font-weight:700; color:#1a1a2e;
    padding:10px 0; border-bottom:3px solid #e94560; margin-bottom:15px;
}
</style>
""", unsafe_allow_html=True)

# ╔══════════════════════════════════════════╗
#   ثوابت
# ╚══════════════════════════════════════════╝
DATA_FILE = "daily_entries.json"
today_str = date.today().strftime("%Y-%m-%d")

# ╔══════════════════════════════════════════╗
#   🔤 نظام التحويل عربي ↔ إنجليزي
# ╚══════════════════════════════════════════╝

# جدول تحويل الحروف العربية → أقرب صوت إنجليزي
ARABIC_TO_LATIN = {
    'ا': 'a', 'أ': 'a', 'إ': 'a', 'آ': 'a',
    'ب': 'b',
    'ت': 't',
    'ث': 'th',
    'ج': 'j',
    'ح': 'h',
    'خ': 'kh',
    'د': 'd',
    'ذ': 'th',
    'ر': 'r',
    'ز': 'z',
    'س': 's',
    'ش': 'sh',
    'ص': 's',
    'ض': 'd',
    'ط': 't',
    'ظ': 'z',
    'ع': 'a',
    'غ': 'gh',
    'ف': 'f',
    'ق': 'q',
    'ك': 'k',
    'ل': 'l',
    'م': 'm',
    'ن': 'n',
    'ه': 'h',
    'و': 'w',
    'ي': 'y', 'ى': 'y',
    'ة': 'h',
    'ء': 'a',
    'ئ': 'y',
    'ؤ': 'w',
    'لا': 'la',
}

# قاموس الأسماء الشائعة عربي → إنجليزي (مخصص للأسماء الفلسطينية والعربية)
NAME_DICT_AR_TO_EN = {
    # أسماء نساء
    "انجود":   ["njwad", "enjwad", "anjoud", "njoud", "engoud"],
    "نجود":    ["njwad", "njoud", "njood"],
    "هويدة":   ["huwayda", "huwaida", "howida"],
    "فاطمة":   ["fatima", "fatema"],
    "مريم":    ["mariam", "maryam"],
    "سارة":    ["sara", "sarah"],
    "نور":     ["nour", "noor", "nur"],
    "لينا":    ["lina", "leena"],
    "رنا":     ["rana"],
    "دينا":    ["dina", "deena"],
    "سمر":     ["samar"],
    "رهف":     ["rahaf"],
    "ريم":     ["reem", "rim"],
    "هند":     ["hind"],
    "غزل":     ["ghazal"],
    "رغد":     ["raghad"],
    "لمى":     ["lama", "luma"],
    "شذى":     ["shatha"],
    "وفاء":    ["wafa", "wafaa"],
    "أسماء":   ["asma", "asmaa"],
    "إيمان":   ["iman", "eiman"],
    "شيماء":   ["shaimaa", "shaima"],
    "منال":    ["manal"],
    "ابتسام":  ["ibtisam"],
    "أميرة":   ["amira", "ameera"],
    "هبة":     ["heba", "hiba"],
    "رشا":     ["rasha"],
    "نادية":   ["nadia", "nadya"],
    "سوسن":    ["sawsan", "suzan"],
    "بسمة":    ["basma", "bassma"],
    "ميسم":    ["maysam"],
    "ملك":     ["malak"],
    "روان":    ["rawan", "rowan"],
    "تالا":    ["tala"],
    "يسرا":    ["yusra", "yosra"],
    "صفاء":    ["safaa", "safa"],
    "جنان":    ["jinan", "janan"],
    "نسرين":   ["nasreen", "nasrin"],
    "حنان":    ["hanan"],
    "سلمى":    ["salma"],
    "زينب":    ["zainab", "zaynab"],
    "ندى":     ["nada", "neda"],
    "شيرين":   ["shirin", "shireen"],

    # أسماء رجال
    "كمال":    ["kamal", "karnal", "kemal"],
    "محمد":    ["mohammad", "mohammed", "muhammad", "mohamad"],
    "احمد":    ["ahmad", "ahmed"],
    "علي":     ["ali"],
    "عمر":     ["omar", "umar"],
    "خالد":    ["khaled", "khalid"],
    "سعيد":    ["saeed", "said", "saied"],
    "حمادة":   ["hamada", "hammad"],
    "حماده":   ["hamada", "hammad"],
    "سليمان":  ["sulaiman", "suleiman", "sliman"],
    "منتصر":   ["muntaser", "montaser"],
    "ياسر":    ["yaser", "yasser"],
    "طارق":    ["tariq", "tarek"],
    "موسى":    ["musa", "moussa"],
    "عيسى":    ["issa", "eissa"],
    "يوسف":    ["yousef", "yusuf", "joseph"],
    "ابراهيم": ["ibrahim", "abraham"],
    "اسماعيل": ["ismail", "ismaeel"],
    "عبدالله": ["abdullah", "abdallah"],
    "عبد الله":["abdullah", "abdallah"],
    "عبدالحميد":["abdulhamid", "abdelhamid"],
    "عبد الحميد":["abdulhamid", "abdelhamid"],
    "خليل":    ["khalil", "halil"],
    "رمضان":   ["ramadan", "ramadhan"],
    "فتحي":    ["fathi", "fathy"],
    "توفيق":   ["tawfiq", "taufiq"],
    "نعيم":    ["naim", "naeem"],
    "صالح":    ["saleh", "salih"],
    "تامر":    ["tamer", "tamir"],
    "مصطفى":   ["mustafa", "mostafa"],
    "ماهر":    ["maher", "mahir"],
    "سالم":    ["salem", "salim"],
    "سلامة":   ["salama"],
    "فايز":    ["fayez", "faiz"],
    "رائد":    ["raed", "raid"],
    "وليد":    ["walid", "waleed"],
    "هاني":    ["hani", "hany"],
    "نبيل":    ["nabil", "nabeel"],
    "جمال":    ["jamal", "gamal"],
    "حسام":    ["hossam", "husam"],
    "عصام":    ["essam", "isam"],
    "بلال":    ["bilal", "belal"],
    "زياد":    ["ziad", "ziyad"],
    "باسل":    ["basel", "basil"],
    "نادر":    ["nader", "nadir"],
    "رامي":    ["rami", "ramy"],
    "عادل":    ["adel", "adil"],
    "جابر":    ["jaber", "jabir"],
    "ماجد":    ["majed", "majid"],
    "حاتم":    ["hatem", "hatim"],
    "اشرف":    ["ashraf"],
    "هيثم":    ["haitham", "haytham"],
    "عمرو":    ["amro", "amr"],
    "كريم":    ["karim", "kareem"],
    "معاذ":    ["muath", "moaz"],
    "انس":     ["anas"],
    "قيس":     ["qais", "qays"],
    "فراس":    ["firas"],
    "لؤي":     ["luay", "louay"],
    "شادي":    ["shadi", "shady"],
    "حازم":    ["hazem", "hazim"],
    "عزيز":    ["aziz"],
    "منير":    ["munir", "monir"],
    "رفيق":    ["rafiq", "rafeeq"],
    "حسن":     ["hassan", "hasan"],
    "حسين":    ["hussein", "husain"],
    "مازن":    ["mazen", "mazin"],
    "يزيد":    ["yazid", "yazeed"],
    "شريف":    ["sharif", "sherif"],

    # ألقاب / عائلات فلسطينية شائعة
    "العبادله":  ["alabadla", "alabadlah", "alabadleh", "aladla"],
    "العبادلة":  ["alabadla", "alabadlah", "alabadleh"],
    "العيادله":  ["alayada", "alayadia", "aliadlh", "alayadleh"],
    "العيادلة":  ["alayada", "alayadia", "alayadleh"],
    "العبدالله": ["alabdla", "alabdlah", "alabdallah"],
    "الاسطل":   ["alastal", "alastel"],
    "الكطوت":   ["alkatout", "alkatut"],
    "العقاد":   ["alaqad", "alakkad"],
    "بركة":     ["baraka", "barka"],
    "المصري":   ["almasri", "almasry", "almasree"],
    "ابو":      ["abu", "abo"],
    "عطيه":     ["atiyah", "atia", "atieh"],
    "عوف":      ["awf", "ouf"],
}

# قاموس عكسي: إنجليزي → عربي
NAME_DICT_EN_TO_AR = {}
for ar, en_list in NAME_DICT_AR_TO_EN.items():
    for en in en_list:
        NAME_DICT_EN_TO_AR[en.lower()] = ar


def arabic_to_phonetic(text: str) -> str:
    """تحويل نص عربي إلى مقابله الصوتي اللاتيني"""
    result = ""
    i = 0
    text = text.strip()
    while i < len(text):
        # جرب حرفين أولاً
        two = text[i:i+2]
        if two in ARABIC_TO_LATIN:
            result += ARABIC_TO_LATIN[two]
            i += 2
        elif text[i] in ARABIC_TO_LATIN:
            result += ARABIC_TO_LATIN[text[i]]
            i += 1
        elif text[i] == ' ':
            result += ' '
            i += 1
        else:
            result += text[i]
            i += 1
    return result.strip()


def normalize_for_match(text: str) -> str:
    """
    توحيد النص للمقارنة:
    - إزالة التشكيل
    - توحيد الأحرف العربية المتشابهة
    - تحويل إلى lowercase
    """
    # إزالة التشكيل
    text = re.sub(r'[\u064B-\u065F\u0670]', '', text)
    text = text.strip().lower()
    # توحيد الأحرف
    for a, b in [("أ","ا"),("إ","ا"),("آ","ا"),("ة","ه"),("ى","ي"),("ك","ك")]:
        text = text.replace(a, b)
    text = re.sub(r'\s+', ' ', text)
    return text


def name_to_search_variants(name: str) -> list[str]:
    """
    توليد كل المتغيرات الممكنة لاسم ما (عربي أو إنجليزي)
    للمقارنة مع الجانب الآخر
    """
    variants = set()
    name_clean = normalize_for_match(name)
    variants.add(name_clean)

    is_arabic = bool(re.search(r'[\u0600-\u06FF]', name))

    if is_arabic:
        # 1. الصوتي المباشر
        phonetic = arabic_to_phonetic(name_clean)
        variants.add(phonetic)

        # 2. البحث في القاموس كلمة كلمة
        words = name_clean.split()
        en_words_options = []
        for word in words:
            word_variants = [word]  # الكلمة الأصلية
            # بحث مباشر
            if word in NAME_DICT_AR_TO_EN:
                word_variants.extend(NAME_DICT_AR_TO_EN[word])
            # بحث بعد التوحيد
            norm_word = normalize_for_match(word)
            if norm_word in NAME_DICT_AR_TO_EN:
                word_variants.extend(NAME_DICT_AR_TO_EN[norm_word])
            # الصوتي
            word_variants.append(arabic_to_phonetic(word))
            en_words_options.append(list(set(word_variants)))

        # 3. توليد جميع التوليفات
        from itertools import product
        for combo in product(*en_words_options):
            variants.add(' '.join(combo))
            # بدون مقاطع القصيرة (ال، ابو)
            filtered = [w for w in combo if len(w) > 2]
            if filtered:
                variants.add(' '.join(filtered))

    else:
        # النص إنجليزي → حوله لعربي
        words = name_clean.split()
        ar_words_options = []
        for word in words:
            word_variants = [word]
            if word in NAME_DICT_EN_TO_AR:
                word_variants.append(NAME_DICT_EN_TO_AR[word])
            ar_words_options.append(list(set(word_variants)))

        from itertools import product
        for combo in product(*ar_words_options):
            variants.add(' '.join(combo))

    return [v for v in variants if v.strip()]


def smart_name_similarity(name_a: str, name_b: str) -> int:
    """
    حساب درجة التشابه الذكية بين اسمين
    يأخذ أعلى درجة من كل المتغيرات الممكنة
    """
    variants_a = name_to_search_variants(name_a)
    variants_b = name_to_search_variants(name_b)

    best_score = 0
    for va in variants_a:
        for vb in variants_b:
            # token_sort_ratio: يرتب الكلمات قبل المقارنة (يتجاهل الترتيب)
            s1 = fuzz.token_sort_ratio(va, vb)
            # partial_ratio: يبحث عن التطابق الجزئي
            s2 = fuzz.partial_ratio(va, vb)
            # token_set_ratio: الأفضل للأسماء المركبة
            s3 = fuzz.token_set_ratio(va, vb)
            score = max(s1, s2, s3)
            if score > best_score:
                best_score = score
            if best_score == 100:
                return 100

    return best_score


def extract_names_from_detail(detail_text: str) -> list[str]:
    """
    استخراج الأسماء من حقل التفاصيل في كشف الحساب
    مثل: "تحويل الى محمد احمد علي"
    """
    names = []

    # نمط: نص عربي بعد "الى" أو "من ... الى"
    patterns = [
        r'الى\s+([\u0600-\u06FF\s]{5,50}?)(?:\s*\d|$|/)',
        r'الي\s+([\u0600-\u06FF\s]{5,50}?)(?:\s*\d|$|/)',
        r'to\s+([a-zA-Z\s]{5,50}?)(?:\s*\d|$|/)',
        # أسماء إنجليزية في كشوف البنوك مثل karnal alabadla/Njwad Alabdla
        r'/([a-zA-Z][a-zA-Z\s]{3,40})/([a-zA-Z][a-zA-Z\s]{3,40})/',
    ]

    for pat in patterns:
        matches = re.findall(pat, detail_text, re.IGNORECASE)
        for m in matches:
            if isinstance(m, tuple):
                names.extend([x.strip() for x in m if x.strip()])
            elif m.strip():
                names.append(m.strip())

    return names


# ╔══════════════════════════════════════════╗
#   دوال التخزين JSON
# ╚══════════════════════════════════════════╝
def load_data() -> dict:
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_data(data: dict):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def get_today_entries() -> list:
    return load_data().get(today_str, [])

def add_entry(name: str, amount: float):
    data = load_data()
    if today_str not in data:
        data[today_str] = []
    data[today_str].append({
        "id":     len(data[today_str]) + 1,
        "name":   name.strip(),
        "amount": amount,
        "time":   datetime.now().strftime("%H:%M:%S")
    })
    save_data(data)

def delete_entry(entry_id: int):
    data = load_data()
    if today_str in data:
        data[today_str] = [e for e in data[today_str] if e["id"] != entry_id]
        save_data(data)

def clear_today():
    data = load_data()
    data[today_str] = []
    save_data(data)

# ╔══════════════════════════════════════════╗
#   استخراج البيانات من الملفات
# ╚══════════════════════════════════════════╝
NAME_KW   = ["اسم","name","الاسم","العميل","client","customer","مستفيد","المستفيد"]
AMOUNT_KW = ["مبلغ","amount","المبلغ","قيمة","value","total","إجمالي","رسوم","الرسوم",
             "المبالغ المستلمة","المبالغ المدفوعة","مستلمة","مدفوعة"]

def clean_amount(val: str):
    val = re.sub(r"[^\d.]", "", str(val))
    try:
        return float(val) if val else None
    except ValueError:
        return None

def find_col_idx(row_values, keywords):
    for i, cell in enumerate(row_values):
        cell_str = str(cell).strip().lower()
        if any(k in cell_str for k in keywords):
            return i
    return None

# ── PDF المتخصص لكشوف البنوك ──
def extract_pdf_bank_statement(file) -> pd.DataFrame:
    """
    استخراج متخصص لكشوف حسابات البنوك الفلسطينية
    يستخرج الأسماء من حقل التفاصيل
    """
    rows = []

    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()

            if tables:
                for table in tables:
                    if not table:
                        continue

                    # ابحث عن صف الرأس
                    header_idx = 0
                    detail_col = amount_received_col = amount_paid_col = None

                    for ri, row in enumerate(table[:5]):
                        if not row:
                            continue
                        row_str = [str(c).strip() if c else "" for c in row]

                        # ابحث عن أعمدة التفاصيل والمبالغ
                        for ci, cell in enumerate(row_str):
                            cell_l = cell.lower()
                            if "تفاصيل" in cell_l or "detail" in cell_l or "البيان" in cell_l:
                                detail_col = ci
                            if "مستلمة" in cell_l or "received" in cell_l or "واردة" in cell_l or "credit" in cell_l:
                                amount_received_col = ci
                            if "مدفوعة" in cell_l or "paid" in cell_l or "صادرة" in cell_l or "debit" in cell_l:
                                amount_paid_col = ci

                        if detail_col is not None:
                            header_idx = ri
                            break

                    # استخراج البيانات
                    for row in table[header_idx + 1:]:
                        if not row:
                            continue
                        row_vals = [str(c).strip() if c else "" for c in row]

                        # استخرج المبلغ (مستلم أو مدفوع)
                        amount = None
                        if amount_received_col is not None and amount_received_col < len(row_vals):
                            amount = clean_amount(row_vals[amount_received_col])
                        if (amount is None or amount == 0) and amount_paid_col is not None and amount_paid_col < len(row_vals):
                            amount = clean_amount(row_vals[amount_paid_col])
                        # إذا لم نجد، ابحث عن أي خلية تحتوي رقم > 0
                        if amount is None or amount == 0:
                            for cell in row_vals:
                                a = clean_amount(cell)
                                if a and a > 0:
                                    amount = a
                                    break

                        # استخرج الاسم من حقل التفاصيل
                        detail_text = ""
                        if detail_col is not None and detail_col < len(row_vals):
                            detail_text = row_vals[detail_col]
                        else:
                            # دمج كل الخلايا كنص
                            detail_text = " ".join(row_vals)

                        # استخراج الأسماء من التفاصيل
                        extracted_names = extract_names_from_detail(detail_text)

                        # إذا ما فيش أسماء من النمط، جرب الأسماء الإنجليزية في النص
                        if not extracted_names:
                            en_names = re.findall(
                                r'\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,4})\b',
                                detail_text
                            )
                            extracted_names = en_names

                        if amount and amount > 0:
                            if extracted_names:
                                for name in extracted_names:
                                    rows.append({"name": name, "amount": amount, "detail": detail_text})
                            else:
                                # احفظ التفاصيل كاملة كـ "اسم" للمقارنة
                                if detail_text and len(detail_text) > 3:
                                    rows.append({"name": detail_text[:80], "amount": amount, "detail": detail_text})

            else:
                # استخراج نصي
                text = page.extract_text() or ""
                for line in text.split("\n"):
                    amount = None
                    amt_match = re.search(r'\b(\d{1,6}(?:\.\d{1,2})?)\b', line)
                    if amt_match:
                        amount = float(amt_match.group(1))

                    names = extract_names_from_detail(line)
                    if not names:
                        en = re.findall(r'\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,3})\b', line)
                        names = en

                    if amount and amount > 0 and names:
                        for name in names:
                            rows.append({"name": name, "amount": amount, "detail": line})

    if not rows:
        return pd.DataFrame(columns=["name","amount","detail"])

    df = pd.DataFrame(rows).drop_duplicates(subset=["name","amount"])
    return df


def extract_excel(file) -> pd.DataFrame:
    rows = []
    xl = pd.ExcelFile(file)
    for sheet in xl.sheet_names:
        raw = xl.parse(sheet, header=None)
        name_col = amount_col = None

        for ri in range(min(15, len(raw))):
            row = raw.iloc[ri].tolist()
            nc = find_col_idx(row, NAME_KW)
            ac = find_col_idx(row, AMOUNT_KW)
            if nc is not None and ac is not None:
                name_col, amount_col = nc, ac
                raw = raw.iloc[ri+1:].reset_index(drop=True)
                break

        if name_col is None:
            name_col, amount_col = 0, 1

        for _, row in raw.iterrows():
            try:
                name = str(row.iloc[name_col]).strip()
                amt  = clean_amount(str(row.iloc[amount_col]))
                if name and name.lower() not in ["nan","none",""] and amt and amt > 0:
                    rows.append({"name": name, "amount": amt, "detail": ""})
            except Exception:
                continue

    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=["name","amount","detail"])


def extract_csv(file) -> pd.DataFrame:
    df = pd.DataFrame()
    for enc in ["utf-8","utf-8-sig","cp1256","latin-1"]:
        try:
            df = pd.read_csv(file, encoding=enc)
            file.seek(0)
            break
        except Exception:
            file.seek(0)

    if df.empty:
        return pd.DataFrame(columns=["name","amount","detail"])

    nc = next((c for c in df.columns if any(k in str(c).lower() for k in NAME_KW)),   df.columns[0])
    ac = next((c for c in df.columns if any(k in str(c).lower() for k in AMOUNT_KW)), df.columns[min(1,len(df.columns)-1)])

    rows = []
    for _, row in df.iterrows():
        name = str(row[nc]).strip()
        amt  = clean_amount(str(row[ac]))
        if name and name.lower() not in ["nan","none",""] and amt and amt > 0:
            rows.append({"name": name, "amount": amt, "detail": ""})
    return pd.DataFrame(rows)


def smart_extract(file, filename: str) -> pd.DataFrame:
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext in ["xlsx","xls"]:
        return extract_excel(file)
    elif ext == "pdf":
        return extract_pdf_bank_statement(file)
    elif ext == "csv":
        return extract_csv(file)
    else:
        try:
            r = extract_excel(file)
            if not r.empty:
                return r
        except Exception:
            pass
        file.seek(0)
        return extract_pdf_bank_statement(file)


# ╔══════════════════════════════════════════╗
#   المطابقة الذكية مع دعم عربي/إنجليزي
# ╚══════════════════════════════════════════╝
def match_entries(
    manual: list,
    file_df: pd.DataFrame,
    name_threshold: int = 70,
    amount_tolerance: float = 0.01
) -> pd.DataFrame:

    results   = []
    file_copy = file_df.copy().reset_index(drop=True)
    file_copy["_used"] = False

    for entry in manual:
        m_name   = entry["name"]
        m_amount = entry["amount"]

        best_score = 0
        best_idx   = -1

        # قارن مع كل سجل في الملف
        for idx, file_row in file_copy[~file_copy["_used"]].iterrows():
            # احسب التشابه الذكي
            score = smart_name_similarity(m_name, file_row["name"])

            # إذا فيه حقل تفاصيل، جرب معه كمان
            if "detail" in file_copy.columns and file_row.get("detail",""):
                detail_names = extract_names_from_detail(str(file_row["detail"]))
                for dn in detail_names:
                    s2 = smart_name_similarity(m_name, dn)
                    if s2 > score:
                        score = s2

            if score > best_score:
                best_score = score
                best_idx   = idx

        if best_score >= name_threshold and best_idx >= 0:
            file_amt = file_copy.at[best_idx, "amount"]
            diff     = round(m_amount - file_amt, 2)
            amt_ok   = abs(diff) <= amount_tolerance * max(m_amount, 1)

            file_copy.at[best_idx, "_used"] = True

            results.append({
                "الاسم (يدوي)":      m_name,
                "المبلغ (يدوي)":     m_amount,
                "الاسم (ملف)":       file_copy.at[best_idx, "name"],
                "المبلغ (ملف)":      file_amt,
                "الفرق":             diff,
                "درجة التشابه":      f"{best_score}%",
                "الحالة":            "✅ مطابق" if amt_ok else "⚠️ اسم مطابق / مبلغ مختلف",
                "_status":           "matched" if amt_ok else "partial"
            })
        else:
            results.append({
                "الاسم (يدوي)":  m_name,
                "المبلغ (يدوي)": m_amount,
                "الاسم (ملف)":   f"أقرب: {file_copy.at[best_idx,'name'] if best_idx>=0 else '—'} ({best_score}%)",
                "المبلغ (ملف)":  0,
                "الفرق":         m_amount,
                "درجة التشابه":  f"{best_score}%",
                "الحالة":        "❌ غير موجود في الملف",
                "_status":       "not_found"
            })

    # ما تبقى في الملف
    for idx, row in file_copy[~file_copy["_used"]].iterrows():
        results.append({
            "الاسم (يدوي)":  "—",
            "المبلغ (يدوي)": 0,
            "الاسم (ملف)":   row["name"],
            "المبلغ (ملف)":  row["amount"],
            "الفرق":         row["amount"],
            "درجة التشابه":  "—",
            "الحالة":        "🔍 موجود في الملف فقط",
            "_status":       "file_only"
        })

    return pd.DataFrame(results)


# ╔══════════════════════════════════════════╗
#   تصدير Excel
# ╚══════════════════════════════════════════╝
def export_excel(match_df: pd.DataFrame, manual: list) -> bytes:
    wb = openpyxl.Workbook()
    STATUS_COLORS = {
        "matched":   "D4EDDA",
        "partial":   "FFF3CD",
        "not_found": "F8D7DA",
        "file_only": "D1ECF1"
    }
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill("solid", fgColor="1A1A2E")
    h_font = Font(color="FFFFFF", bold=True, size=11)

    ws1 = wb.active
    ws1.title = "تقرير المطابقة"
    ws1.sheet_view.rightToLeft = True

    cols = ["الاسم (يدوي)","المبلغ (يدوي)","الاسم (ملف)","المبلغ (ملف)","الفرق","درجة التشابه","الحالة"]
    for ci, h in enumerate(cols, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.fill = h_fill; c.font = h_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for ri, row in match_df[cols].iterrows():
        color = STATUS_COLORS.get(match_df.at[ri, "_status"], "FFFFFF")
        fill  = PatternFill("solid", fgColor=color)
        for ci, val in enumerate(row, 1):
            c = ws1.cell(row=ri+2, column=ci, value=val)
            c.fill = fill
            c.alignment = Alignment(horizontal="center")
            c.border = border

    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 26

    ws2 = wb.create_sheet("الإدخال اليومي")
    ws2.sheet_view.rightToLeft = True
    for ci, h in enumerate(["م","الاسم","المبلغ","الوقت"], 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.fill = h_fill; c.font = h_font
    for i, e in enumerate(manual, 2):
        ws2.cell(row=i, column=1, value=e["id"])
        ws2.cell(row=i, column=2, value=e["name"])
        ws2.cell(row=i, column=3, value=e["amount"])
        ws2.cell(row=i, column=4, value=e.get("time",""))

    ws3 = wb.create_sheet("الملخص")
    ws3.sheet_view.rightToLeft = True
    total_manual = sum(e["amount"] for e in manual)
    stats = [
        ("التاريخ",                  today_str),
        ("إجمالي المدخلات اليدوية",  total_manual),
        ("✅ مطابق تماماً",          len(match_df[match_df["_status"]=="matched"])),
        ("⚠️ مبلغ مختلف",           len(match_df[match_df["_status"]=="partial"])),
        ("❌ غير موجود",            len(match_df[match_df["_status"]=="not_found"])),
        ("🔍 في الملف فقط",         len(match_df[match_df["_status"]=="file_only"])),
    ]
    for ri, (k, v) in enumerate(stats, 1):
        ws3.cell(row=ri, column=1, value=k).font = Font(bold=True)
        ws3.cell(row=ri, column=2, value=v)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ╔══════════════════════════════════════════╗
#   الواجهة الرئيسية
# ╚══════════════════════════════════════════╝
st.markdown(f"""
<div class="header-box">
  <h1>📊 نظام المطابقة اليومية</h1>
  <p>يدعم المطابقة بين الأسماء العربية والإنجليزية تلقائياً</p>
  <p style="color:#e94560;">📅 {today_str}</p>
</div>
""", unsafe_allow_html=True)

for key in ["file_df","match_result"]:
    if key not in st.session_state:
        st.session_state[key] = None

entries = get_today_entries()

# ── شريط جانبي ──
with st.sidebar:
    st.markdown('<div class="section-title">➕ إدخال يومي</div>', unsafe_allow_html=True)

    with st.form("entry_form", clear_on_submit=True):
        name   = st.text_input("👤 الاسم", placeholder="مثال: انجود العبادله")
        amount = st.number_input("💰 المبلغ", min_value=0.0, step=0.01, format="%.2f")
        if st.form_submit_button("➕ إضافة"):
            if name.strip() and amount > 0:
                add_entry(name, amount)
                st.success(f"✅ {name} — {amount:,.2f}")
                st.session_state.match_result = None
                st.rerun()
            else:
                st.error("⚠️ أدخل الاسم والمبلغ")

    st.divider()

    # 🔍 اختبار التشابه
    with st.expander("🧪 اختبر تشابه اسمين"):
        test_a = st.text_input("الاسم الأول", placeholder="انجود العبادله")
        test_b = st.text_input("الاسم الثاني", placeholder="Njwad Alabdla")
        if st.button("احسب التشابه") and test_a and test_b:
            score = smart_name_similarity(test_a, test_b)
            st.metric("درجة التشابه", f"{score}%")
            with st.expander("المتغيرات المولّدة"):
                st.write("**الاسم الأول:**", name_to_search_variants(test_a))
                st.write("**الاسم الثاني:**", name_to_search_variants(test_b))

    st.divider()

    st.markdown('<div class="section-title">📂 رفع الكشف</div>', unsafe_allow_html=True)
    uploaded = st.file_uploader("Excel / PDF / CSV", type=["xlsx","xls","pdf","csv"])

    if uploaded:
        with st.spinner("⏳ استخراج البيانات..."):
            df = smart_extract(uploaded, uploaded.name)
        if not df.empty:
            st.session_state.file_df = df
            st.success(f"✅ {len(df)} سجل")
            with st.expander("👁️ معاينة"):
                st.dataframe(df[["name","amount"]].head(15), use_container_width=True)
        else:
            st.error("❌ لم يُعثر على بيانات")

    st.divider()

    st.markdown('<div class="section-title">⚙️ إعدادات</div>', unsafe_allow_html=True)
    threshold = st.slider("🎯 حساسية تشابه الاسم", 50, 95, 70)
    tolerance = st.slider("💱 هامش فرق المبلغ %", 0, 10, 1) / 100

    st.divider()
    c1, c2 = st.columns(2)
    with c1:
        if st.button("🔄 مطابقة"):
            if not entries:
                st.error("لا يوجد مدخلات!")
            elif st.session_state.file_df is None:
                st.error("لم يُرفع ملف!")
            else:
                with st.spinner("⏳ جاري المطابقة..."):
                    st.session_state.match_result = match_entries(
                        entries,
                        st.session_state.file_df,
                        name_threshold=threshold,
                        amount_tolerance=tolerance
                    )
                st.success("✅ اكتملت!")
    with c2:
        if st.button("🗑️ مسح"):
            clear_today()
            st.session_state.match_result = None
            st.session_state.file_df      = None
            st.rerun()

# ── تبويبات ──
tab1, tab2, tab3 = st.tabs(["📋 المدخلات اليومية","🔍 نتائج المطابقة","📅 السجل التاريخي"])

with tab1:
    entries = get_today_entries()
    total = sum(e["amount"] for e in entries)
    count = len(entries)
    c1,c2,c3,c4 = st.columns(4)
    c1.markdown(f'<div class="metric-card card-blue"><div class="metric-value">{count}</div><div class="metric-label">عدد الإدخالات</div></div>', unsafe_allow_html=True)
    c2.markdown(f'<div class="metric-card card-green"><div class="metric-value">{total:,.2f}</div><div class="metric-label">الإجمالي</div></div>', unsafe_allow_html=True)
    c3.markdown(f'<div class="metric-card card-orange"><div class="metric-value">{total/count:,.2f if count else 0}</div><div class="metric-label">المتوسط</div></div>', unsafe_allow_html=True)
    c4.markdown(f'<div class="metric-card card-red"><div class="metric-value">{today_str}</div><div class="metric-label">التاريخ</div></div>', unsafe_allow_html=True)

    st.markdown("---")
    if entries:
        df_show = pd.DataFrame(entries)[["id","name","amount","time"]]
        df_show.columns = ["م","الاسم","المبلغ","الوقت"]
        st.dataframe(df_show, use_container_width=True, hide_index=True)
        del_id = st.number_input("رقم الإدخال للحذف", min_value=1, step=1)
        if st.button("🗑️ حذف"):
            delete_entry(int(del_id))
            st.rerun()
    else:
        st.info("📭 لا توجد مدخلات اليوم")

with tab2:
    if st.session_state.match_result is not None:
        mr = st.session_state.match_result
        matched   = len(mr[mr["_status"]=="matched"])
        partial   = len(mr[mr["_status"]=="partial"])
        not_found = len(mr[mr["_status"]=="not_found"])
        file_only = len(mr[mr["_status"]=="file_only"])
        total_r   = len(mr)
        pct       = round(matched/total_r*100,1) if total_r else 0

        c1,c2,c3,c4,c5 = st.columns(5)
        c1.markdown(f'<div class="metric-card card-green"><div class="metric-value">{matched}</div><div class="metric-label">✅ مطابق</div></div>',    unsafe_allow_html=True)
        c2.markdown(f'<div class="metric-card card-orange"><div class="metric-value">{partial}</div><div class="metric-label">⚠️ فرق مبلغ</div></div>', unsafe_allow_html=True)
        c3.markdown(f'<div class="metric-card card-red"><div class="metric-value">{not_found}</div><div class="metric-label">❌ غير موجود</div></div>',  unsafe_allow_html=True)
        c4.markdown(f'<div class="metric-card card-blue"><div class="metric-value">{file_only}</div><div class="metric-label">🔍 ملف فقط</div></div>',   unsafe_allow_html=True)
        c5.markdown(f'<div class="metric-card card-green"><div class="metric-value">{pct}%</div><div class="metric-label">نسبة المطابقة</div></div>',    unsafe_allow_html=True)

        st.progress(pct/100)
        st.markdown("---")

        filter_opts = st.multiselect(
            "🔎 فلترة",
            ["✅ مطابق","⚠️ اسم مطابق / مبلغ مختلف","❌ غير موجود في الملف","🔍 موجود في الملف فقط"],
            default=["✅ مطابق","⚠️ اسم مطابق / مبلغ مختلف","❌ غير موجود في الملف","🔍 موجود في الملف فقط"]
        )
        display_cols = ["الاسم (يدوي)","المبلغ (يدوي)","الاسم (ملف)","المبلغ (ملف)","الفرق","درجة التشابه","الحالة"]
        filtered = mr[mr["الحالة"].apply(lambda x: any(s in x for s in filter_opts))]
        st.dataframe(filtered[display_cols], use_container_width=True, hide_index=True)

        entries_now = get_today_entries()
        if entries_now:
            xl = export_excel(mr, entries_now)
            st.download_button("📥 تحميل التقرير Excel", xl,
                               f"تقرير_{today_str}.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info("⏳ ارفع ملفاً واضغط 🔄 مطابقة")

with tab3:
    st.markdown('<div class="section-title">📅 السجل التاريخي</div>', unsafe_allow_html=True)
    all_data = load_data()
    if all_data:
        for day in sorted(all_data.keys(), reverse=True):
            day_ent   = all_data[day]
            day_total = sum(e["amount"] for e in day_ent)
            with st.expander(f"📅 {day} — {len(day_ent)} إدخال | الإجمالي: {day_total:,.2f}"):
                if day_ent:
                    df_d = pd.DataFrame(day_ent)[["id","name","amount","time"]]
                    df_d.columns = ["م","الاسم","المبلغ","الوقت"]
                    st.dataframe(df_d, use_container_width=True, hide_index=True)
    else:
        st.info("📭 لا يوجد سجل بعد")